import ttkbootstrap as ttk
from sympy.strategies.core import switch
from ttkbootstrap.constants import *
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import openpyxl
from datetime import datetime
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ------------------- FASE DA INTERFACE GRAFICA -------------------------


dados = {"nome_planilha": None}  # Dicionário para armazenar o valor


def enviar_selecao():

    selecoes = [opcao for opcao, var in toggles.items() if var.get()]
    # Criando a planilha
    workbook = openpyxl.Workbook()
    try:
        del workbook['Sheet']
    except KeyError:
        print("Não achei essa planilha")

    for i in selecoes:
        pagina = workbook.create_sheet(i)
        pagina['A1'].value = 'Nome'
        pagina['B1'].value = 'Preço'
        pagina['C1'].value = 'Estoque'
        pagina['D1'].value = 'Avaliação'

    dados["nome_planilha"] = 'ToScrapLivros_BETA' + str(datetime.today().date()) + '.xlsx'
    workbook.save(dados["nome_planilha"])
    root.quit()


# Criando a janela principal
root = ttk.Window(themename="darkly")
root.title("Escolha de Livros")
root.geometry("350x250")
root.resizable(False, False)

# Título
titulo = ttk.Label(root, text="Escolha um gênero literário", font=("Arial", 14, "bold"))
titulo.pack(pady=10)

# Frame para organizar os toggles em colunas
frame_toggles = ttk.Frame(root)
frame_toggles.pack()


# Lista de opções
opcoes = ["Travel", "Mystery", "Historical-Fiction", "Sequential-Art", "Classics", "Philosophy", "Romance"]
toggles = {opcao: ttk.BooleanVar() for opcao in opcoes}

# Criando os toggles em duas colunas
for i, (opcao, var) in enumerate(toggles.items()):
    ttk.Checkbutton(frame_toggles, text=opcao, variable=var, bootstyle="success-round-toggle").grid(row=i//2, column=i%2, padx=10, pady=5, sticky="w")

# Botão para enviar a seleção
btn_enviar = ttk.Button(root, text="Confirmar Seleção", bootstyle="primary", command=enviar_selecao)
btn_enviar.pack(pady=15)

# Rodando a interface
root.mainloop()




# ------------------- FASE DE WEB SCRAPING -----------------------

livros_categorias = {
    'Travel': 2,
    'Mystery': 3,
    'Historical-Fiction': 4,
    'Sequential-Art': 5,
    'Classics': 6,
    'Philosophy': 7,
    'Romance': 8
}

# AGORA A PARTE DE WEB SCRAPING
service = Service()
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)

# 1. Primeiro devo acessar o arquivo .xlsx e pegar os nomes das páginas
planilha = openpyxl.load_workbook(dados["nome_planilha"])
pagina = planilha.sheetnames # Isso é basicamente uma lista com o nome de todas as páginas

# 2. Agora construo cada URL separadamente conforme o nome da página
for i in range(len(pagina)):
    url = "https://books.toscrape.com/catalogue/category/books/"+pagina[i].lower()+"_"+str(livros_categorias[pagina[i]])+"/index.html"
    driver.get(url)

    # Listas para armazenar dados
    nomes = []
    precos = []
    estoques = []
    avaliacoes = []
    links = []

    # Coletar os links de todos os livros
    while True:
        # Esperar até que os livros carreguem (espera 5 segundos)
        elementos = WebDriverWait(driver, 5).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'h3 a'))
        )
        # Adicionar os links dos livros à lista
        for elemento in elementos:
            links.append(elemento.get_attribute('href'))

        # Verificar se há um botão "next" para carregar mais produtos
        try:
            botao = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//a[contains(text(),'next')]"))
            )
            # Clicar no botão "next" para carregar a próxima página de livros
            ActionChains(driver).move_to_element(botao).click(botao).perform()
            print("Botão 'next' clicado! Carregando mais produtos...")
        except:
            print("Botão 'next' não encontrado. Todos os produtos foram carregados!")
            break  # Sai do loop quando não encontrar o botão "next"

    # Agora, para cada link, abrir o livro e coletar os dados
    for link in links:
        driver.get(link)
        try:
            nome = driver.find_element(By.CSS_SELECTOR, 'h1').text
            preco = driver.find_elements(By.CLASS_NAME, 'price_color')[0].text
            estoque_texto = driver.find_element(By.CSS_SELECTOR, '.instock.availability').text
            estoque = int(estoque_texto.split('(')[-1].split()[0])  # Extração segura
            rating = driver.find_element(By.CSS_SELECTOR, ".star-rating").get_attribute("class").split(' ')[-1]
            class_to_number = {
                "One": 1,
                "Two": 2,
                "Three": 3,
                "Four": 4,
                "Five": 5
            }
            avaliacao = class_to_number.get(rating, 0)

            # Adicionar os dados na lista
            nomes.append(nome)
            precos.append(preco)
            estoques.append(estoque)
            avaliacoes.append(avaliacao)
        except Exception as e:
            print(f"Erro ao coletar dados do livro {link}: {e}")

    # Agora com as listas prontas, vamos adicionar as informações no Excel
    sheet = planilha[pagina[i]]
    for j in range(len(nomes)):
        sheet.append([nomes[j], precos[j], estoques[j], avaliacoes[j]])

    print(f"Dados da categoria {pagina[i]} coletados com sucesso!")

    planilha.save(dados["nome_planilha"])

# Fechar o driver do Selenium
driver.quit()
